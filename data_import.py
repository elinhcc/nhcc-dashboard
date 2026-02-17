"""Excel import: read provider spreadsheet, parse fax numbers, convert to Vonage format, populate DB."""
import re
import openpyxl
from datetime import datetime
from database import get_connection, init_db, add_practice, add_provider
from utils import backup_excel, load_config, categorize_location


def parse_fax_number(contact_info: str) -> str:
    """Extract fax number from contact information text."""
    if not contact_info:
        return ""
    # Look for fax patterns
    patterns = [
        r'[Ff][Aa][Xx]\s*(?:[Nn]umber)?\s*:?\s*\(?\s*(\d{3})\s*\)?\s*[-.\s]*(\d{3})\s*[-.\s]*(\d{4})',
        r'[Ff][Aa][Xx]\s*:?\s*(\d{3})\s*[-.\s]*(\d{3})\s*[-.\s]*(\d{4})',
    ]
    for pattern in patterns:
        match = re.search(pattern, contact_info)
        if match:
            groups = match.groups()
            return f"{groups[0]}-{groups[1]}-{groups[2]}"
    return ""


def parse_phone_number(contact_info: str) -> str:
    """Extract phone number from contact information text."""
    if not contact_info:
        return ""
    # Remove fax lines to avoid picking up fax as phone
    lines = contact_info.split("\n")
    non_fax_lines = [l for l in lines if "fax" not in l.lower()]
    text = "\n".join(non_fax_lines) if non_fax_lines else contact_info

    patterns = [
        r'(?:[Pp]hone|[Pp][Hh]|[Tt]el|[Oo]ffice)\s*:?\s*\(?\s*(\d{3})\s*\)?\s*[-.\s]*(\d{3})\s*[-.\s]*(\d{4})',
        r'\((\d{3})\)\s*(\d{3})\s*[-.\s]*(\d{4})',
        r'(\d{3})\s*[-.\s](\d{3})\s*[-.\s](\d{4})',
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            groups = match.groups()
            return f"{groups[0]}-{groups[1]}-{groups[2]}"
    return ""


def convert_fax_to_vonage_email(fax_number) -> str:
    """Convert a fax number to Vonage email format: 1(XXX)XXXXXXX@fax.vonagebusiness.com."""
    if not fax_number or str(fax_number).strip() == '':
        return ""
    cleaned = re.sub(r'[^0-9]', '', str(fax_number))
    if len(cleaned) < 10:
        return ""
    if len(cleaned) == 11 and cleaned[0] == '1':
        cleaned = cleaned[1:]
    elif len(cleaned) > 10:
        cleaned = cleaned[-10:]
    area_code = cleaned[:3]
    local_number = cleaned[3:]
    return f"1({area_code}){local_number}@fax.vonagebusiness.com"


def fax_to_vonage_email(fax_number: str, domain: str = "fax.vonagebusiness.com") -> str:
    """Legacy wrapper — calls convert_fax_to_vonage_email."""
    return convert_fax_to_vonage_email(fax_number)


def parse_providers(providers_text: str) -> list:
    """Parse provider names from the Providers column."""
    if not providers_text:
        return []
    providers = []
    lines = providers_text.strip().split("\n")
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Skip category headers like "PA's:", "NPs:", etc.
        if re.match(r'^[A-Z]+[\'\u2019]?s?\s*:?\s*$', line):
            continue
        # Skip numbered prefixes
        line = re.sub(r'^\d+[\.\)]\s*', '', line)
        # Remove leading role prefixes like "N-", "Dr."
        line = re.sub(r'^N-', '', line)
        if line and len(line) > 1:
            providers.append(line.strip())
    return providers


def extract_zip(address: str) -> str:
    """Extract zip code from address."""
    if not address:
        return ""
    match = re.search(r'\b(\d{5})(?:-\d{4})?\b', address)
    return match.group(1) if match else ""


def _detect_columns(ws) -> dict:
    """Read header row and map column names to indices.

    Falls back to positional defaults if headers aren't recognized.
    """
    header_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val:
            header_map[str(val).strip().lower()] = col

    # Map known header names to our field names
    aliases = {
        "practice_name": ["practice name", "practice", "name", "office name", "clinic name"],
        "address": ["address", "location", "office address"],
        "contact_info": ["contact information", "contact info", "phone/fax", "phone fax", "contact"],
        "providers": ["providers", "provider", "doctors", "physician"],
        "details": ["details", "notes", "detail"],
        "next_followup": ["next follow up", "next followup", "follow up", "followup"],
        "last_contact": ["last contact date", "last contact", "last contacted"],
    }

    col_map = {}
    for field, names in aliases.items():
        for name in names:
            if name in header_map:
                col_map[field] = header_map[name]
                break

    # Fallback: positional mapping based on Excel structure (Col 1-based)
    col_map.setdefault("practice_name", 1)
    col_map.setdefault("address", 2)
    col_map.setdefault("contact_info", 3)
    col_map.setdefault("providers", 4)
    col_map.setdefault("details", 5)
    col_map.setdefault("next_followup", 6)
    col_map.setdefault("last_contact", 8)

    return col_map


def import_excel(excel_path: str = None) -> dict:
    """Import the Excel file into the database. Returns import stats."""
    config = load_config()
    if not excel_path:
        excel_path = config["excel_path"]

    # Backup first
    backup_path = backup_excel()

    # Initialize database
    init_db()

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Full List"]

    # Detect column layout from headers
    col_map = _detect_columns(ws)

    stats = {
        "practices_imported": 0,
        "providers_imported": 0,
        "fax_numbers_found": 0,
        "rows_processed": 0,
        "skipped_rows": 0,
        "backup_path": backup_path,
        "column_map": {k: v for k, v in col_map.items()},
    }

    current_practice_name = None

    for row in range(2, ws.max_row + 1):
        practice_name = ws.cell(row, col_map["practice_name"]).value
        address = ws.cell(row, col_map["address"]).value
        contact_info = ws.cell(row, col_map["contact_info"]).value
        providers_text = ws.cell(row, col_map["providers"]).value
        details = ws.cell(row, col_map["details"]).value
        next_followup = ws.cell(row, col_map["next_followup"]).value
        last_contact = ws.cell(row, col_map["last_contact"]).value

        # Skip completely empty rows
        if not practice_name and not address and not contact_info and not providers_text:
            stats["skipped_rows"] += 1
            continue

        # Handle sub-rows (no practice name = continuation of previous)
        if not practice_name and current_practice_name:
            practice_name = current_practice_name + " (Branch)"
        elif practice_name:
            current_practice_name = practice_name
        else:
            stats["skipped_rows"] += 1
            continue

        stats["rows_processed"] += 1

        # Parse contact info
        contact_str = str(contact_info) if contact_info else ""
        fax = parse_fax_number(contact_str)
        phone = parse_phone_number(contact_str)
        vonage_email = convert_fax_to_vonage_email(fax) if fax else ""
        zip_code = extract_zip(str(address) if address else "")
        location = categorize_location(str(address) if address else "")

        if fax:
            stats["fax_numbers_found"] += 1

        # Build notes from details and follow-up
        notes_parts = []
        if details:
            notes_parts.append(str(details))
        if next_followup:
            notes_parts.append(f"Next follow-up: {next_followup}")
        notes = "\n".join(notes_parts)

        # Format last contact date
        last_contact_str = None
        if last_contact:
            if isinstance(last_contact, datetime):
                last_contact_str = last_contact.isoformat()
            else:
                last_contact_str = str(last_contact)

        # Add practice
        practice_data = {
            "name": str(practice_name).strip(),
            "address": str(address).strip() if address else "",
            "zip_code": zip_code,
            "location_category": location,
            "phone": phone,
            "fax": fax,
            "fax_vonage_email": vonage_email,
            "contact_person": "",
            "notes": notes,
            "status": "Active",
        }
        practice_id = add_practice(practice_data)
        stats["practices_imported"] += 1

        # If we have a last contact date, add it to contact log
        if last_contact_str:
            from database import add_contact_log
            add_contact_log({
                "practice_id": practice_id,
                "contact_type": "Other",
                "contact_date": last_contact_str,
                "team_member": "Robbie",
                "notes": "Imported from Excel",
            })

        # Parse and add providers
        provider_names = parse_providers(str(providers_text) if providers_text else "")
        for pname in provider_names:
            add_provider({
                "name": pname,
                "practice_id": practice_id,
                "status": "Active",
            })
            stats["providers_imported"] += 1

    wb.close()

    # Post-import fixup: ensure every practice with a fax number has a vonage email
    try:
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, fax FROM practices "
            "WHERE fax IS NOT NULL AND fax != '' "
            "AND (fax_vonage_email IS NULL OR fax_vonage_email = '')"
        ).fetchall()
        fixed = 0
        for r in rows:
            vonage = convert_fax_to_vonage_email(r["fax"])
            if vonage:
                conn.execute(
                    "UPDATE practices SET fax_vonage_email=? WHERE id=?",
                    (vonage, r["id"]),
                )
                fixed += 1
        conn.commit()
        conn.close()
        if fixed:
            stats["vonage_emails_fixed"] = fixed
    except Exception:
        pass

    # Auto-save database to GitHub after successful import
    try:
        from database_persistence import save_database_to_github
        save_database_to_github()
    except Exception:
        pass

    return stats


def get_import_status():
    """Check if data has been imported."""
    import os
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "providers.db")
    if not os.path.exists(db_path):
        return False
    try:
        conn = get_connection()
        count = conn.execute("SELECT COUNT(*) FROM practices").fetchone()[0]
        conn.close()
        return count > 0
    except Exception:
        return False
